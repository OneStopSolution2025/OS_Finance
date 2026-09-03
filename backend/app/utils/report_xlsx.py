import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.config import settings

XLSX_REPORTS_DIR = os.path.join(settings.LOCAL_STORAGE_PATH, "reports")
os.makedirs(XLSX_REPORTS_DIR, exist_ok=True)

CHARCOAL_FILL = PatternFill(start_color="183B66", end_color="183B66", fill_type="solid")
HEADER_FONT = Font(color="64A844", bold=True, size=11)
DANGER_FILL = PatternFill(start_color="FBE4E1", end_color="FBE4E1", fill_type="solid")
DANGER_FONT = Font(color="C0392B", bold=False, size=10)
LATE_FONT = Font(color="C0392B", bold=True, size=10)
GROUP_LABEL_HEADERS = {"day": "Date", "week": "Week", "month": "Month", "employee": "Employee", "branch": "Branch"}


def generate_breakdown_xlsx(tenant_name: str, group_by: str, rows: list[dict], outstanding_rows: list[dict] | None = None) -> str:
    """
    One row per actual payment collected — never a rolled-up total — followed
    by a highlighted sheet section listing every overdue, still-unpaid
    installment, including exactly which member within a group hasn't paid.
    """
    outstanding_rows = outstanding_rows or []
    wb = Workbook()
    ws = wb.active
    ws.title = f"{group_by.capitalize()} Breakdown"[:31]

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Udhayam MFI — {tenant_name} — {group_by.capitalize()}-wise Collections"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {datetime.utcnow().strftime('%d %b %Y, %I:%M %p UTC')} — {len(rows)} paid, {len(outstanding_rows)} outstanding"
    ws["A2"].font = Font(italic=True, size=9, color="888888")

    group_col_label = GROUP_LABEL_HEADERS.get(group_by, "Group")

    # ---- Section 1: Collected ----
    row_cursor = 4
    ws.cell(row=row_cursor, column=1, value=f"COLLECTED ({len(rows)} payments)").font = Font(bold=True, size=12, color="183B66")
    row_cursor += 1
    header_row = row_cursor
    headers = [group_col_label, "Payment Date", "Paid By", "Type", "Group", "Employee", "Loan #", "Receipt #", "Status", "Amount (₹)"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = CHARCOAL_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    total_amount = 0
    row_cursor = header_row + 1
    for row in rows:
        ws.cell(row=row_cursor, column=1, value=row["group_label"])
        ws.cell(row=row_cursor, column=2, value=row["date"])
        ws.cell(row=row_cursor, column=3, value=row["payer_name"])
        ws.cell(row=row_cursor, column=4, value=row["payer_type"])
        ws.cell(row=row_cursor, column=5, value=row["group_name"])
        ws.cell(row=row_cursor, column=6, value=row["employee_name"])
        ws.cell(row=row_cursor, column=7, value=row["loan_number"])
        ws.cell(row=row_cursor, column=8, value=row["receipt_number"])
        status_cell = ws.cell(row=row_cursor, column=9, value=row["status"])
        if "Late" in row["status"]:
            status_cell.font = LATE_FONT
        ws.cell(row=row_cursor, column=10, value=row["amount"]).number_format = "#,##0.00"
        total_amount += row["amount"]
        row_cursor += 1

    ws.cell(row=row_cursor, column=1, value="TOTAL COLLECTED").font = Font(bold=True)
    ws.cell(row=row_cursor, column=8, value=f"{len(rows)} payments").font = Font(bold=True)
    ws.cell(row=row_cursor, column=10, value=round(total_amount, 2)).font = Font(bold=True)
    ws.cell(row=row_cursor, column=10).number_format = "#,##0.00"
    row_cursor += 3

    # ---- Section 2: Outstanding / Not Paid — highlighted ----
    ws.cell(row=row_cursor, column=1, value=f"⚠ OUTSTANDING / NOT PAID ({len(outstanding_rows)} overdue)").font = Font(bold=True, size=12, color="C0392B")
    row_cursor += 1
    out_header_row = row_cursor
    out_headers = [group_col_label, "Due Date", "Owes", "Type", "Group", "Employee", "Loan #", "Days Late", "Amount Due (₹)"]
    for col, text in enumerate(out_headers, start=1):
        cell = ws.cell(row=out_header_row, column=col, value=text)
        cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")

    total_outstanding = 0
    row_cursor = out_header_row + 1
    for row in outstanding_rows:
        for col, value in enumerate([
            row["group_label"], row["due_date"], row["payer_name"], row["payer_type"],
            row["group_name"], row["employee_name"], row["loan_number"], f"{row['days_overdue']}d",
        ], start=1):
            cell = ws.cell(row=row_cursor, column=col, value=value)
            cell.fill = DANGER_FILL
            cell.font = DANGER_FONT
        amount_cell = ws.cell(row=row_cursor, column=9, value=row["amount_due"])
        amount_cell.number_format = "#,##0.00"
        amount_cell.fill = DANGER_FILL
        amount_cell.font = DANGER_FONT
        total_outstanding += row["amount_due"]
        row_cursor += 1

    ws.cell(row=row_cursor, column=1, value="TOTAL OUTSTANDING").font = Font(bold=True, color="C0392B")
    ws.cell(row=row_cursor, column=7, value=f"{len(outstanding_rows)} unpaid").font = Font(bold=True, color="C0392B")
    ws.cell(row=row_cursor, column=9, value=round(total_outstanding, 2)).font = Font(bold=True, color="C0392B")
    ws.cell(row=row_cursor, column=9).number_format = "#,##0.00"

    widths = [14, 13, 20, 13, 20, 16, 12, 14, 12, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    filename = f"breakdown-{group_by}-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.xlsx"
    file_path = os.path.join(XLSX_REPORTS_DIR, filename)
    wb.save(file_path)
    return file_path
